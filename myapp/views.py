import cv2
import os
import numpy as np
from django.shortcuts import render, redirect, get_object_or_404
from myapp.models import Performance, User,Student,Parent,PROFESSOR,Comment,Profile
from myapp.forms import CommentForm
import face_recognition
from django.contrib.auth import authenticate, login as auth_login
from django.http import HttpResponse, HttpResponseRedirect

from django.contrib import messages

def register(request):
    show_identifier_field = False
    show_identifier_field1 = False
    show_identifier_field3 = False


    if request.method == 'POST':
        selected_option = request.POST.get('role')
        if selected_option == 'Student':
            show_identifier_field = True
        if selected_option == 'Parent':
            show_identifier_field1 = True
        if selected_option == 'Professor':
            show_identifier_field3 = True
        # Open the default camera
        cap = cv2.VideoCapture(0)

        # Get a face encoding and save the face image for the new user
        print("Please look at the camera for a few seconds...")
        while True:
            ret, frame = cap.read()
            if not ret:
                print("Error: Could not read frame from camera")
                break

            # Find all the faces in the frame
            face_locations = face_recognition.face_locations(frame)

            if len(face_locations) == 1:
                # Get the face encoding and face image for the first face found
                face_encoding = face_recognition.face_encodings(frame, face_locations)[0]
                face_image = frame[face_locations[0][0]:face_locations[0][2], face_locations[0][3]:face_locations[0][1]]

                # Save the face image to file
                file_name = request.POST['username'] + ".jpg"
                file_path = os.path.join('faces', file_name)
                cv2.imwrite(file_path, face_image)
                print("Face saved successfully!")
                break

        # Release the camera
        cap.release()

        # Save the user in the User model
        username = request.POST['username']
        user = User(role=selected_option, username=username, face_encoding=face_encoding.tobytes())
        if show_identifier_field:
            student_id = request.POST.get('student_id')
            level1=request.POST.get('level')
            student, created = Student.objects.get_or_create(id=student_id,level=level1)
            user.student_id = student.id
            user.level=student.level
            user.save()
        if show_identifier_field1:
            print("1")
            student_id = request.POST.get('student_id1')
            student = get_object_or_404(Student, id=student_id)
            if student:
               parents = Parent.objects.filter(id=student_id)
               if parents.exists():
                  parent = parents.first()  # Choose one parent if multiple are found
               else:
                  parent = Parent.objects.create()  # Create a new parent if none are found
                  user.parent_id = parent.id 
                  parent.students.add(student)
                  user.student_id=student.id
                  user.save()
                  return render(request, 'registration/login.html', {'student': student})
            else:
                # If the student_id does not exist, show an error message or handle it accordingly
                messages.error(request, 'Invalid student_id')
                return redirect('register')
        if show_identifier_field3:
                cin = request.POST.get('cin')
                professor, created = PROFESSOR.objects.get_or_create(id=cin)
                user.professor_id=professor.id
                students = Student.objects.all()
                user.save()
                return render(request, 'registration/login.html', {'students': students})

                

          
    return render(request, 'registration/register.html', {'show_identifier_field': show_identifier_field})
from django.contrib.auth.decorators import login_required

@login_required
def dashboard(request,user):
    user = request.user

    if user.role == 'Student':
        return redirect('classification')
    elif user.role == 'Parent':
        return redirect('parent_profile_view')
    elif user.role == 'Professor':
        return redirect('add_dataset')
    else:
        # Handle other roles or default case
        return render(request, 'default_dashboard.html')

def login(request):
    if request.method == 'POST':
        # Open the default camera
        cap = cv2.VideoCapture(0)

        # Get a face encoding for the user trying to login
        print("Please look at the camera for a few seconds...")
        while True:
            ret, frame = cap.read()
            if not ret:
                print("Error: Could not read frame from camera")
                break

            # Find all the faces in the frame
            face_locations = face_recognition.face_locations(frame)

            if len(face_locations) == 1:
                # Get the face encoding for the first face found
                face_encoding = face_recognition.face_encodings(frame, face_locations)[0]
                face_image = frame[face_locations[0][0]:face_locations[0][2], face_locations[0][3]:face_locations[0][1]]
                break

        # Release the camera
        cap.release()

        # Compare the face encoding to the face encodings of all the users in the "faces" folder
        match_found = False
        for filename in os.listdir('faces'):
            if not filename.endswith('.jpg'):
                continue

            encodings = face_recognition.face_encodings(face_recognition.load_image_file(os.path.join('faces', filename)))
        
                
        match = face_recognition.compare_faces([encodings][0], face_encoding)

        if match[0]:
                    # Login the user
                    #username = os.path.splitext(filename)[0]
                    username = request.POST.get('username')
                    print("1",username)
                    user = User.objects.get(username=username)
                    print(user)

                    if user is not None:
                        auth_login(request, user)
                        print(request, f'You have been logged in {user.username}!')
                        return dashboard(request,user)                        
                    else:
                        messages.error(request, 'Authentication error.')
                        return redirect('login')

        match_found = True

        if not match_found:
            # If no match found, show error message
            messages.error(request, 'Could not recognize your face.')
            return redirect('login')

    return render(request, 'registration/login.html')
from django.contrib.auth import logout
from django.shortcuts import redirect

 # replace 'home' with the URL name of your homepage
def logout_view(request):
    # Check if the user is authenticated
    if request.user.is_authenticated:
        # Do something if authenticated
        logout(request)
        pass 
    else:
        # Check if the user has already seen the welcome message
        if 'welcome_message' not in request.session:
            # Store a message in the session to indicate that the user has seen it
            request.session['welcome_message'] = True
            # Render the welcome page
            return redirect('login')
    
    # Render the main page
    return redirect('login')
from django.contrib.auth.decorators import login_required


@login_required
def create_comment(request):
    if request.user.is_authenticated:
        user_id = request.user.id
        source_user = get_object_or_404(User, id=user_id)
    else:
        return redirect('login')
    
    if request.method == 'POST':
        destination_user_id = request.POST.get('destination_user_id')
        destination_user = get_object_or_404(User, id=destination_user_id)
        content = request.POST.get('content')
        form = CommentForm(request.POST, request.FILES)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.source_user = source_user
            comment.destination_user=destination_user
            comment.content=content
            comment.save()
            return redirect('comment')
    else:
        if source_user.role == 'Student':
            professors = User.objects.filter(role='Professor')
            form = CommentForm(professors=professors)
        elif source_user.role == 'Professor':
            students = User.objects.filter(role='Student')
            form = CommentForm(students=students)
        else:
            return redirect('login')  # Handle invalid role case
    
    return render(request, 'registration/cretecomment.html', {'form': form})



@login_required
def comment_list(request):
    user = request.user
    role = user.role

    if role == 'Student':
        sent_comments = Comment.objects.filter(source_user=user)
        received_comments = Comment.objects.filter(destination_user=user)
    elif role == 'Professor':
        sent_comments = Comment.objects.filter(source_user=user)
        received_comments = Comment.objects.filter(destination_user=user)
    else:
        sent_comments = Comment.objects.none()
        received_comments = Comment.objects.none()

    return render(request, 'registration/comment.html', {'sent_comments': sent_comments, 'received_comments': received_comments})
def profile_view(request, user_id):
    # Get the user object
    user = get_object_or_404(User, id=user_id)

    try:
        # Try to get the associated profile
        profile = Profile.objects.get(user=user)
    except Profile.DoesNotExist:
        # If profile doesn't exist, create a new one
        profile = Profile.objects.create(user=user)  # Customize this based on your Profile model fields and requirements

    context = {
        'user': user,
        'profile': profile,
    }

    return render(request, 'registration/profile.html', context)
def student_profile_view(request, student_id):
    # Retrieve the student object
    student = get_object_or_404(Student, id=student_id)
    user.parent_id = parent.id 
    parent.students.add(student)
    user.student_id=student.id
    # Make sure the logged-in user is the parent of the student
    if request.user != student.parent:
        # If the logged-in user is not the parent, handle it appropriately (e.g., redirect to an error page)
        return HttpResponse('Unauthorized access')

    # Render the student profile page
    context = {
        'student': student,
    }
    return render(request, 'registration/student_profile.html', context)



from django.shortcuts import render, get_object_or_404
from myapp.models import Parent
def parent_profile_view(request):

    parent = get_object_or_404(User, parent_id=request.user.parent_id)

    students =  get_object_or_404(Student,id=request.user.student_id)  # Retrieve associated students
    
   
    performance = Performance.objects.filter(student_id=students)
    
    comment = Comment.objects.filter(destination_user=request.user)
    form = CommentForm()

    context = {
        'parent': parent,
        'performance': performance,
         'student': students,
         'comment':comment,
        'comment_form': form,

    }

    return render(request, 'registration/parent_profile.html', context)























































from django import forms
from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from nbconvert import HTMLExporter
import nbformat
import os
import pandas as pd
from transformers import pipeline
from django.shortcuts import render
import io
import nbformat
import openpyxl
import random
import pyttsx3

 


 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 
 

    
    
    
    
    
    
    #TraductionArabic
    


def TraductionArabic(request):
    # Load the data
    df = pd.read_excel('data/miniData.xlsx')
    # Select a random question
    question = df.sample()["                                                                                                             Question"].iloc[0]
    # Initialize the BERT-based translation pipeline for English to Arabic
    translator = pipeline("translation_en_to_ar", model="Helsinki-NLP/opus-mt-en-ar")
    # Translate the question to Arabic
    translated = translator(question)
    context = {
        'question': question,
        'translated_question': translated[0]["translation_text"],
    }
    return render(request, 'TraductionArabic.html', context)









# TraductionFrancais


def TraductionFrancais(request):
    # Load the data
    df = pd.read_excel('data/miniData.xlsx')
    # Select a random question
    question = df.sample()["                                                                                                             Question"].iloc[0]
    # Initialize the BERT-based translation pipeline for English to Arabic
    translator = pipeline("translation_en_to_ar", model='Helsinki-NLP/opus-mt-en-fr')

    # Translate the question to Arabic
    translated = translator(question)
    context = {
        'question': question,
        'translated_question': translated[0]["translation_text"],
    }
    return render(request, 'TraductionFrancais.html', context)



















# generate sound 

import pyttsx3
import openpyxl
import random
# import the time module
import time

from django.shortcuts import render



def TextToSpeech(request):
    # create a pyttsx3 object with an English accent
    engine = pyttsx3.init() 
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[1].id)  # 1 is the index of the English voice

    # set the reading speed to a slower rate
    rate = engine.getProperty('rate')
    engine.setProperty('rate', int(rate * 0.75))  # adjust the rate as needed

 

    if request.method == 'POST':
        if 'speak' in request.POST:
            # get the selected question from the form
            question = request.POST.get('question')

            # use the pyttsx3 engine to speak the selected question aloud
            engine.say(question)
            engine.runAndWait()

            # return the selected question as a context variable to be displayed in the template
            context = {'question': question}
            return render(request, 'TextToSpeech.html', context)
        
        elif 'next' in request.POST:
            # read the Excel file containing the questions
            workbook = openpyxl.load_workbook('data/miniData.xlsx')
            sheet = workbook['Sheet1']

            # get the questions
            questions = []
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                questions.append(row[0])

            # randomly select another question from the list
            question = random.choice(questions)

            # return the selected question as a context variable to be displayed in the template
            context = {'question': question}
            return render(request, 'TextToSpeech.html', context)
    else:
        # read the Excel file containing the questions
        workbook = openpyxl.load_workbook('data/miniData.xlsx')
        sheet = workbook['Sheet1']

        # get the questions
        questions = []
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            questions.append(row[0])
 
        # randomly select a question from the list
        question = random.choice(questions)

        # return the selected question as a context variable to be displayed in the template
        context = {'question': question}
        return render(request, 'TextToSpeech.html', context)


# define a form for selecting a question
class QuestionForm(forms.Form):
    question = forms.CharField(widget=forms.HiddenInput())
    generate_sound = forms.BooleanField(label='Generate Sound', required=False)






















    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    

def result(request):
    answer = request.POST.get('answer')
    

    question = request.session.get('question')

    solution = request.session.get('solution')
    image_data = request.FILES['image']
    print(image_data)

    user_answer = answer or image_data 
    if image_data:   
        vision=extract_text(request, solution,image_data)

    print(solution)
        # Check if the user's answer is correct
    if user_answer == str(solution):  # Convert solution to string for comparison
            score = 10  # Correct answer, assign 10 as the score
    else:
            score = 0  # Wrong answer, assign 0 as the score

        # Create a new Performance object and store the score
    student = Student.objects.get(id=request.user.student_id)

        # Create a Performance object and save it to the database
    performance = Performance(student_id=request.user.student_id, question=question,solution_provided=user_answer)
    performance.student = student

    performance.date = date.today()  # Set the date of the performance
    performance.score = score  # Assign the calculated score
    performance.save()
    if user_answer and solution:
        user_answer = user_answer.lower().strip()
       
        similarity_score = fuzz.partial_ratio(user_answer, str(solution))
     
        if similarity_score >= 80:
            context = {'result': 'Congratulations! Your answer is correct ', 'is_correct': True ,'image': correct_img}
        else:
            context = {'result': f"Sorry, your answer '{user_answer}' is incorrect. The correct answer is {solution}.", 'is_correct': False ,'image': incorrect_img}
    else:
        context = {'result': 'Please enter an answer', 'image': None}
    return render(request, 'result.html', context or vision)




 
    








    # formule
import pandas as pd
import spacy
import re
import numpy as np
from django.shortcuts import render
def formula_view(correct_answer,question):
    nlp = spacy.load("en_core_web_sm")

    doc = nlp(question)
    question_keywords = [token.text for token in doc if not token.is_stop and token.pos_ in ['NUM', 'NOUN', 'VERB', 'ADJ']]
    keywords_str = ' '.join(question_keywords)

    def predict_operator(operands, answer):
        # Predict operator
        operators = ["+", "-", "*", "/"]
        results = [eval(f"{operands[0]} {operator} {operands[1]}") for operator in operators]
        closest_result = min(results, key=lambda x: abs(x - answer))
        answer = correct_answer

        closest_operator = f"{operands[0]}{operators[np.argmin([abs(result - answer) for result in results])]}{operands[1]}"

        return closest_operator

    def predict_operator1(operands, answer):
        operators = ["+", "-", "*", "/"]
        operand1 = int(operands[0])
        operand2 = int(operands[1])
        answer = correct_answer

        results = [eval(f"{operand1} {operator} {operand2}") for operator in operators]
        closest_result = min(results, key=lambda x: abs(x - answer))
        closest_operator = f"{operators[np.argmin([abs(result - answer) for result in results])]}"

        return closest_operator

    operands = re.findall(r"[\d.]+", keywords_str)
   

    answer = correct_answer

    operator = predict_operator(operands, answer)
    formula = predict_operator1(operands, answer)

    context = {
        'question': question,
        'operator': operator,
        'formula': formula,
    }

    return  formula, operator























from django.shortcuts import render
from django.http import HttpResponse
import nltk
nltk.download('stopwords')

import re
from nltk.corpus import stopwords
from PIL import Image
import requests
from io import BytesIO
import pandas as pd

def generate_image(request):
    path_to_excel_file = 'miniData.xlsx'

    df = pd.read_excel(path_to_excel_file, engine="openpyxl")
    text2 = df.sample(n=1).iloc[0]['                                                                                                             Question']
    print("Question: ", text2)
    text=text2

    question_match = re.search("(?<=, ).*(?=\?)", text)
    if question_match:
     question = question_match.group(0)
     text = text.replace(question, '')
    else:
     question = None


    # Split the text into sentences
    sentences = re.split("(?<!\w\,\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s", text)

    # Define the stop words
    stop_words = set(stopwords.words('english'))

    # Create a list of image objects
    images = []

    # Input your OpenAI API key
    API_KEY = "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"
    # Iterate through each sentence and generate images for the filtered words
    for sentence in sentences:
        # Filter the words
        words = sentence.split()
        filtered_words = [word for word in words if word.lower() not in stop_words and word != "?"]
        print(filtered_words)

        # Generate an image for each filtered word
        for word in filtered_words:
            response = requests.post(
                "https://api.openai.com/v1/images/generations",
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {API_KEY}"
                },
                json={
                    "model": "image-alpha-001",
                    "prompt": f"Generate an image of \"{word}\".",
                    "num_images": 1,
                    "size": "256x256",
                    "response_format": "url"
                })

            # Check if there was an error in the response
            if response.status_code != 200:
                print(f"Error: {response.status_code}")
                print(response.json())
            else:
                # Get the image URL from the API response
                data = response.json()
                if "data" in data and len(data["data"]) > 0:
                    image_url = data["data"][0]["url"]
                    # Download the image from the URL
                    image_data = requests.get(image_url).content
                    # Create a PIL image object from the image data
                    image = Image.open(BytesIO(image_data))
                    # Append the image object to the list
                    images.append(image)
                else:
                    print("Error: Empty response data")

    # Calculate the total width and maximum height of the images
    total_width = sum(image.width for image in images)
    max_height = max(image.height for image in images)

    # Create a new image object with the calculated size
    result_image = Image.new(images[0].mode, (total_width, max_height))

    # Paste each image next to the previous one
    x_offset = 0
    for image in images:
        result_image.paste(image, (x_offset, 0))
        x_offset += image.width
        
        # Save the result image to a file   
    static_path = os.path.join(settings.STATICFILES_DIRS[0], 'result.png')

    result_image.save(static_path)
     # Render the response with the result image path
    result_image_path = os.path.join(settings.STATIC_URL, 'result.png')
    # Render
    context = {
              'text2': text2,
              'result_image_path':  result_image_path ,
              }
 
    return render(request, 'image.html', context)

















    
import random
from django.shortcuts import render
from transformers import pipeline

def generate_random_question():
    # Load the data
    df = pd.read_excel('data/miniData.xlsx')
    # Select a random row
    random_row = df.sample(n=1).iloc[0]
    # Get the question and answer from the random row
    question = random_row["                                                                                                             Question"]
    answer = random_row["Answer"]
    return question, answer







def generate_sound(question):
    # Generate sound from the question
    sound = TextToSpeech(question)
    return sound




def translate_question(question):
    # Translate the question to different languages
    translator_ar = pipeline("translation_en_to_ar", model="Helsinki-NLP/opus-mt-en-ar")
    translator_fr = pipeline("translation_en_to_fr", model="Helsinki-NLP/opus-mt-en-fr")

    translation_ar = translator_ar(question)
    translation_fr = translator_fr(question)

    return translation_ar[0]["translation_text"], translation_fr[0]["translation_text"]







from django.shortcuts import render
import cv2
import pytesseract 
import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from django.views.decorators.csrf import csrf_exempt




@csrf_exempt
def extract_text(request,question,image_data):
    
    


    

# Create a CountVectorizer object
    vectorizer = CountVectorizer()
    question = str(question)
# Fit and transform the 'Answer' column of the DataFrame using the CountVectorizer object
    X = vectorizer.fit_transform([question])
    if request.method == 'POST':
        # Get the uploaded image from the request
        
        # Load the image
        image = cv2.imdecode(np.frombuffer(image_data, np.uint8), cv2.IMREAD_COLOR)

        # Convert the image to grayscale
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Apply adaptive thresholding to the grayscale image
        thresh_image = cv2.adaptiveThreshold(gray_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 7221, -1.5)

        # Remove noise and smooth the image
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        thresh_image = cv2.morphologyEx(thresh_image, cv2.MORPH_CLOSE, kernel)

        # Find contours in the threshold image
        contours, _ = cv2.findContours(thresh_image, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        # Check that contours were found
        if len(contours) > 0:
            # Find the contour with the largest area, which is likely to be the text region
            largest_contour = max(contours, key=cv2.contourArea)

            # Draw a bounding box around the largest contour
            x, y, w, h = cv2.boundingRect(largest_contour)
            cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)

            # Extract the text from the bounding box using OCR
            text_box = gray_image[y:y+h, x:x+w]
            text = pytesseract.image_to_string(text_box, lang='eng', config=r'--psm 6 --oem 1 ')
        else:
            text = 'No text found.'

        # Convert the extracted text to a numpy array and transform it using the CountVectorizer object
        text_vector = vectorizer.transform(np.array([text]))

        # Compute the cosine similarity between the extracted text and each answer in the dataset
        scores = cosine_similarity(text_vector, X)
        print(text)
        # Compute the majority vote among the scores
        if max(scores)[0] >= 0.7:
            result = "The image contains a 'right answer'"
        else:
            result = "The image does not contain a 'right answer'"

        context = {
            'image': image,
            'text': text,
            'result': result,
        }

        return render(request, 'result.html', context)

    else:
        return render(request, 'result.html',context)




























    # Classification 
    

import os
import random
import re

from django.shortcuts import render
from PIL import Image
from fuzzywuzzy import fuzz
import pandas as pd

# Load the dataset

# Set the path for the images
good_image_path = os.path.abspath("good.png")
bad_image_path = os.path.abspath("bad.jpg")

# Set the desired width of the image
desired_width = 200

# Load the images
correct_image = Image.open(good_image_path)
incorrect_image = Image.open(bad_image_path)

# Calculate the proportional height of the image
width, height = correct_image.size
proportional_height = int((desired_width / float(width)) * float(height))

# Resize the images
correct_img = correct_image.resize((desired_width, proportional_height))
incorrect_img = incorrect_image.resize((desired_width, proportional_height))
from datetime import date

@login_required
def classification(request):
    
    
    question, correct_answer = generate_random_question()
    
    
    
    
    
 
    # Extract the solution from the correct answer using regular expressions
    pattern = r'\d+'
    matches = re.findall(pattern, str(correct_answer))
    if matches:
        solution = int(matches[-1])


    # image =generate_image(request, question)

    TextToSpeech(request)

   # Translate the question
    translation_ar, translation_fr = translate_question(question)
 
  


      # Generate a formula
    formula, operator = formula_view(correct_answer,question)
    
    text=question
    question_match = re.search("(?<=, ).*(?=\?)", text)
    if question_match:
     question1 = question_match.group(0)
     text = text.replace(question1, '')
    else:
     question1 = None



    # Split the text into sentences
    sentences = re.split("(?<!\w\,\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s", text)

    # Define the stop words
    stop_words = set(stopwords.words('english'))

    # Create a list of image objects
    images = []

    # Input your OpenAI API key
    API_KEY = "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"
    # Iterate through each sentence and generate images for the filtered words
    for sentence in sentences:
        # Filter the words
        words = sentence.split()
        filtered_words = [word for word in words if word.lower() not in stop_words and word != "?"]
        print(filtered_words)

        # Generate an image for each filtered word
        for word in filtered_words:
            response = requests.post(
                "https://api.openai.com/v1/images/generations",
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {API_KEY}"
                },
                json={
                    "model": "image-alpha-001",
                    "prompt": f"Generate an image of \"{word}\".",
                    "num_images": 1,
                    "size": "256x256",
                    "response_format": "url"
                })

            # Check if there was an error in the response
            if response.status_code != 200:
                print(f"Error: {response.status_code}")
                print(response.json())
            else:
                # Get the image URL from the API response
                data = response.json()
                if "data" in data and len(data["data"]) > 0:
                    image_url = data["data"][0]["url"]
                    # Download the image from the URL
                    image_data = requests.get(image_url).content
                    # Create a PIL image object from the image data
                    image = Image.open(BytesIO(image_data))
                    # Append the image object to the list
                    images.append(image)
                else:
                    print("Error: Empty response data")

    # Calculate the total width and maximum height of the images
    total_width = sum(image.width for image in images)
    max_height = max(image.height for image in images)

    # Create a new image object with the calculated size
    result_image = Image.new(images[0].mode, (total_width, max_height))

    # Paste each image next to the previous one
    x_offset = 0
    for image in images:
        result_image.paste(image, (x_offset, 0))
        x_offset += image.width
        
        # Save the result image to a file   
    static_path = os.path.join(settings.STATICFILES_DIRS[0], 'result.png')

    result_image.save(static_path)
     # Render the response with the result image path
    result_image_path = os.path.join(settings.STATIC_URL, 'result.png')

   



    context = {'question': question , 
                'question1': question1 , 
               'translation_ar': translation_ar,
               'translation_fr': translation_fr,
               'formula': formula,
               'operator': operator,
               'result_image_path':result_image_path,
              
             }
    
    request.session['solution'] = solution
    request.session['question'] = question

    print(solution)
    
    
    return render(request, 'classification.html', context=context)


# views.py

from django.shortcuts import render, redirect
from .forms import DatasetForm
import pandas as pd
@login_required
def add_dataset(request):
    if request.method == 'POST':
        form = DatasetForm(request.POST)
        if form.is_valid():
            question = form.cleaned_data['question']
            answer = form.cleaned_data['answer']
            level = form.cleaned_data['level']

            # Load the dataset from Excel
            dataset = pd.read_excel('data/miniData.xlsx')

            # Create a new row with the form data
            new_row = {'                                                                                                             Question': question, 'Answer': answer, 'level': level}
            dataset = dataset.append(new_row, ignore_index=True)

            # Save the updated dataset back to Excel
            dataset.to_excel('data/miniData.xlsx', index=False)

            return redirect('add_dataset')
    else:
        form = DatasetForm()

    return render(request, 'registration/student_profile.html', {'form': form})
